from openpyxl import load_workbook
from typing import List, Dict

def write_metrics_to_excel(
    file_path: str,
    model_name: str,
    mode: str,  # "clear" | "RAG" | "LORA" | "FULL FINETUNE"
    metrics: Dict[str, float],
    gpu_name = None,
    path = None,
    quant = None,
    test_path = None
):
    """
    metrics = {
        "perplexity": float,
        "blue": float,
        "tok/j": float,
        "lat": float,
        "tok/s": float
    }
    """
    wb = load_workbook(file_path)
    ws = wb["Лист1"]
    
    # Нормализуем режим и определяем начальную колонку для записи метрик
    mode_norm = mode.strip().upper()
    mode_columns = {
        "CLEAR": 6,
        "RAG": 11,
        "LORA": 16,
        "FULL FINETUNE": 21,
        "FULL": 21
    }
    
    if mode_norm not in mode_columns:
        raise ValueError(f"Unknown mode: {mode}")
    
    start_col = mode_columns[mode_norm]
    
    # Поиск существующей строки с таким же model_name и gpu_name
    target_row = None
    #первая и вторая строка – заголовки, данные начинаются со 3-й
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == model_name and ws.cell(row=row, column=2).value == gpu_name and ws.cell(row=row, column=3).value == path and ws.cell(row=row, column=4).value == quant and ws.cell(row=row, column=5).value == test_path:

            target_row = row
            break
    
    # Если строка не найдена, добавляем новую
    if target_row is None:
        target_row = ws.max_row + 1
        ws.cell(row=target_row, column=1).value = model_name
        ws.cell(row=target_row, column=2).value = gpu_name
        ws.cell(row=target_row, column=3).value = path
        ws.cell(row=target_row, column=4).value = quant
        ws.cell(row=target_row, column=5).value = test_path
    
    # Записываем метрики в соответствующие колонки
    ws.cell(row=target_row, column=start_col).value = metrics["perplexity"]
    ws.cell(row=target_row, column=start_col + 1).value = metrics["blue"]
    ws.cell(row=target_row, column=start_col + 2).value = metrics["tok/j"]
    ws.cell(row=target_row, column=start_col + 3).value = metrics["lat"]
    ws.cell(row=target_row, column=start_col + 4).value = metrics["tok/s"]
    
    wb.save(file_path)